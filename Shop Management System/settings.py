from tkinter import *
import os
import openpyxl as xl

a=os.path.join(os.sys.path[0],'Records')
b=os.path.join(os.sys.path[0],'img')

def deftheme():
    try:
                file_path=a+'\\settings.xlsx'
                wb=xl.load_workbook(file_path)
                sheet=wb['Sheet']
                
    except FileNotFoundError:
                print('aafa')                           
                file_name=' '
                file_name='settings.xlsx'
                wb=xl.Workbook()
                file_path=a+"\\"+file_name
                wb.save(file_path)
                sheet=wb['Sheet']
                sheet.cell(1,1).value='black'
                sheet.cell(2,1).value='red'
                sheet.cell(3,1).value='blue'
                c=b+'\\drk1.png'
                d=b+'\\drk2.png'
                e=b+'\\drk3.png'
                f=b+'\\drks4.png'
                sheet.cell(4,1).value=c
                sheet.cell(5,1).value=d
                sheet.cell(6,1).value=e
                sheet.cell(7,1).value=f
                sheet.cell(1,2).value=1
                wb.save(file_path)
                wb.save(file_path)

                    
def theme():
    root=Tk()
    root.configure(background='black')
    root.title("Shop Management")   
    root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))
    Label(root,text='THEMES',fg="lime",font=("arial baltic",15)).place(x=200,y=70) 
    Label(root,text='LIGHT',fg="White",font=("arial baltic",15)).place(x=0,y=150) 
    Label(root,text='DARK',fg="lime",font=("arial baltic",15)).place(x=0,y=200) 
    Label(root,text='SpiderMan',fg="red",font=("arial baltic",15)).place(x=0,y=250) 
    Label(root,text='IronMan',fg="red",font=("arial baltic",15)).place(x=0,y=300) 
    Label(root,text='Captain',fg="blue",font=("arial baltic",15)).place(x=0,y=350) 
    Label(root,text='Hulk',fg="green",font=("arial baltic",15)).place(x=0,y=400) 
    Button(root,text="LIGHT",fg="lime",font=("arial baltic",15),bg="white",width=root.winfo_screenwidth(),command=lambda:lgt_mode()).place(x=200,y=150) 
    Button(root,text="DARK",fg="lime",font=("arial baltic",15),bg="grey",width=root.winfo_screenwidth(),command=lambda:derk_mode()).place(x=200,y=200) 
    Button(root,text="Spider man",fg="blue",font=("arial baltic",15),bg="red",width=root.winfo_screenwidth(),command=lambda:spd_mode()).place(x=200,y=250) 
    Button(root,text="Iron Man",fg="yellow",font=("arial baltic",15),bg="red",width=root.winfo_screenwidth(),command=lambda:im_mode()).place(x=200,y=300) 
    Button(root,text="captain america",fg="white",font=("arial baltic",15),bg="blue",width=root.winfo_screenwidth(),command=lambda:ca_mode()).place(x=200,y=350) 
    Button(root,text="Hulk",fg="blue",font=("arial baltic",15),bg="dark green",width=root.winfo_screenwidth(),command=lambda:hulk_mode()).place(x=200,y=400) 
    
   
    def derk_mode():
        
        file_path=a+'\\settings.xlsx'
        wb=xl.load_workbook(file_path)
        sheet=wb['Sheet']
        sheet.cell(1,1).value='white'
        sheet.cell(2,1).value='black'
        sheet.cell(3,1).value='grey'
        c=b+'\\drk1.png'
        d=b+'\\drk2.png'
        e=b+'\\drk3.png'
        f=b+'\\drk4.png'
        
        sheet.cell(4,1).value=c
        sheet.cell(5,1).value=d
        sheet.cell(6,1).value=e
        sheet.cell(7,1).value=f
        wb.save(file_path)
        
        wb.save(file_path)
        
    
    def lgt_mode():
        
        file_path=a+'\\settings.xlsx'
        wb=xl.load_workbook(file_path)
        sheet=wb['Sheet']
        sheet.cell(1,1).value='white'
        sheet.cell(2,1).value='blue'
        sheet.cell(3,1).value='grey'
        c=b+'\\wht1.png'
        d=b+'\\wht2.png'
        e=b+'\\wht3.png'
        f=b+'\\wht4.png'
        
        sheet.cell(4,1).value=c
        sheet.cell(5,1).value=d
        sheet.cell(6,1).value=e
        sheet.cell(7,1).value=f
        wb.save(file_path)
        
        
        wb.save(file_path)
        
    
    def spd_mode():
        
        file_path=a+'\\settings.xlsx'
        wb=xl.load_workbook(file_path)
        sheet=wb['Sheet']
        sheet.cell(1,1).value='blue'
        sheet.cell(2,1).value='red'
        sheet.cell(3,1).value='black'
        c=b+'\\spd.png'
        d=b+'\\spd2.png'
        e=b+'\\spd3.png'
        f=b+'\\spd4.png'
        
        sheet.cell(4,1).value=c
        sheet.cell(5,1).value=d
        sheet.cell(6,1).value=e
        sheet.cell(7,1).value=f
        wb.save(file_path)
        
    
    def im_mode():
        
        file_path=a+'\\settings.xlsx'
        wb=xl.load_workbook(file_path)
        sheet=wb['Sheet']
        sheet.cell(1,1).value='yellow'
        sheet.cell(2,1).value='red'
        sheet.cell(3,1).value='light blue'
        sheet.cell(4,1).value='red'
        
        
        c=b+'\\im1.png'
        d=b+'\\im2.png'
        e=b+'\\im3.png'
        f=b+'\\im4.png'
        
        sheet.cell(4,1).value=c
        sheet.cell(5,1).value=d
        sheet.cell(6,1).value=e
        sheet.cell(7,1).value=f
        wb.save(file_path)        
    
    
    def ca_mode():
        
        file_path=a+'\\settings.xlsx'
        wb=xl.load_workbook(file_path)
        sheet=wb['Sheet']
        sheet.cell(1,1).value='white'
        sheet.cell(2,1).value='blue'
        sheet.cell(3,1).value='red'
        sheet.cell(4,1).value='blue'
        c=b+'\\ca1.png'
        d=b+'\\ca2.png'
        e=b+'\\ca3.png'
        f=b+'\\ca4.png'
        
        sheet.cell(4,1).value=c
        sheet.cell(5,1).value=d
        sheet.cell(6,1).value=e
        sheet.cell(7,1).value=f
        wb.save(file_path)
        
    
    def hulk_mode():
        
        file_path=a+'\\settings.xlsx'
        wb=xl.load_workbook(file_path)
        sheet=wb['Sheet']
        sheet.cell(1,1).value='blue'
        sheet.cell(2,1).value='green'
        sheet.cell(3,1).value='black'
        sheet.cell(4,1).value='green'
        c=b+'\\hlk1.png'
        d=b+'\\hlk2.png'
        e=b+'\\hk3.png'
        f=b+'\\hlk4.png'
        
        sheet.cell(4,1).value=c
        sheet.cell(5,1).value=d
        sheet.cell(6,1).value=e
        sheet.cell(7,1).value=f
        wb.save(file_path)
    root.mainloop()
    
               
