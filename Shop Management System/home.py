# pip install openpyxl
# pip install pyttsx3

from tkinter import *
from tkinter import Menu
import data2
import data1
import data3
import login
import pyttsx3
import os
import openpyxl as xl
import settings as stng
from tkinter import messagebox
import webbrowser as vb

engine = pyttsx3.init()


s=os.path.join(os.sys.path[0],'Records')
p=os.path.join(os.sys.path[0],'img')
print(p)

stng.deftheme()

file_path=s+'\\settings.xlsx'
wb=xl.load_workbook(file_path)
sheet=wb['Sheet']
frg=sheet.cell(1,1).value
bkg=sheet.cell(2,1).value
brg=sheet.cell(3,1).value
bkimg=sheet.cell(4,1).value
bkimg2=sheet.cell(5,1).value
bkimg3=sheet.cell(6,1).value
bkimg4=sheet.cell(7,1).value
sound=sheet.cell(1,2).value
engine=pyttsx3.init('espeak')
voices=engine.getProperty('voices')
engine.setProperty('voice',voices[sound].id)

def ajex(sayit):     
       
     
        engine.say(sayit)
        engine.runAndWait()
   
def ajax(root,sayit):
     Button(root,text= "Ajex",bg=brg,padx=30,fg=frg,font="elephant 12 bold",borderwidth=1,relief=SUNKEN,command=lambda:ajex(sayit)).place(x=20,y=20)


    
def home(root1='none'):
    if root1!='none':
        root1.destroy()        
    root=Tk()
    
    image_path = os.path.join(p, 'strt.png').replace('\\', '/')
    photo0 = PhotoImage(file=image_path)
    #photo0=PhotoImage(file=p+'\\strt.png')
    Tphoto=Label(root,image=photo0)
    Tphoto.place(x=00,y=400)
   

    '''photo=PhotoImage(file=bkimg)
    Tphoto=Label(root,image=photo,bg=bkg)
    Tphoto.pack()
    '''
    
    image_path = os.path.join(p, 'gl.png').replace('\\', '/')
    photo2 = PhotoImage(file=image_path)
    #photo2=PhotoImage(file=p+'\\gl.png')
    Tphoto=Button(root,image=photo2,command=lambda:vb.open_new_tab('http://gurukulkhurai.org/gyanodaya/'))
    Tphoto.place(x=1160,y=600)
   
    image_path = os.path.join(p, 'ig.png').replace('\\', '/')
    photo3 = PhotoImage(file=image_path)
    Tphoto=Button(root,image=photo3,command=lambda:vb.open_new_tab('https://www.instagram.com/aadarshrivastava/'))
    Tphoto.place(x=1080,y=600)
    
    image_path = os.path.join(p, 'fb.png').replace('\\', '/')
    photo4 = PhotoImage(file=image_path)
    #photo4=PhotoImage(file=p+'\\fb.png')
    Tphoto=Button(root,image=photo4,command=lambda:vb.open_new_tab(' https://m.facebook.com/aadarsh.shrivastava.752?ref=bookmarks'))
    Tphoto.place(x=1000,y=600)
    
    image_path = os.path.join(p, 'YT.png').replace('\\', '/')
    photo5 = PhotoImage(file=image_path)
    # photo5=PhotoImage(file=p+'\\YT.png')
    Tphoto=Button(root,image=photo5,command=lambda:vb.open_new_tab('https://www.youtube.com/channel/UC_7RhrC1nipjZTanSKoEVEA'))
    Tphoto.place(x=900,y=600)
    
    menuinator(root)
    
    Button(root,image=photo0,fg=frg,font=("arial baltic",20),bg=bkg,command=lambda:data3.edit_sheet()).place(x=600,y=300)  
    Button(root,text="EXIT",fg=frg,font=("arial baltic",20),bg=bkg,command=lambda:root.destroy()).place(x=600,y=600)  
    ajax(root,'hi, welcome to  your shop management panel, i am ajax ,your assistant and friend')
  #  root.overrideredirect(True)
    root.mainloop()


def set_win(root):
 
    root.title("Shop Management")   
    root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))

    root.overrideredirect(True)
    

    
def donothing():
    pass    


def menuinator(root):
    menubar=Menu(root,fg=frg,bg=bkg,font='19',relief='sunken')
    filemenu=Menu(menubar,fg=frg,bg=bkg,font='19',relief='sunken',tearoff=0)
    filemenu.add_command(label="Menu", command=lambda:secwin(root))
    filemenu.add_command(label="Stocks ", command=lambda:stockwin(root))
    filemenu.add_command(label="Sells", command=lambda:sellwin(root))
    filemenu.add_command(label="home", command=lambda:home(root))
    filemenu.add_separator()
    filemenu.add_command(label="Exit", command=root.quit)
    menubar.add_cascade(label="File", menu=filemenu)

    helpmenu = Menu(menubar,fg=frg,bg=bkg,font='19',relief='sunken', tearoff=0)
    helpmenu.add_command(label="Help Index", command=donothing)
    helpmenu.add_command(label="About...", command=lambda:ajex('sorry,  it is still under construction'))
    menubar.add_cascade(label="Help", menu=helpmenu)

    loginmenu= Menu(menubar,fg=frg,bg=bkg,font='19',relief='sunken', tearoff=0)
    loginmenu.add_command(label="Login", command=lambda:login.login_form())
    loginmenu.add_command(label="Sign in", command=lambda:login.signin_form())
    menubar.add_cascade(label="login", menu=loginmenu)

    
    setmenu= Menu(menubar,fg=frg,bg=bkg,font='19',relief='sunken', tearoff=0)
    setmenu.add_command(label="Themes", command=lambda:stng.theme())
    
   
    menubar.add_cascade(label="Settings", menu=setmenu)
    
    
    lgoutmenu= Menu(menubar,fg=frg,bg=bkg,font='19',relief='sunken', tearoff=0)
    lgoutmenu.add_command(label="Logout", command=lambda:login.logout())
    menubar.add_cascade(label="Logout", menu=lgoutmenu)
    
    viewmenu = Menu(menubar,fg=frg,bg=bkg,font='19',relief='sunken', tearoff=0)
    viewmenu.add_command(label="View current file", command=lambda:data3.show())
    viewmenu.add_command(label="View searched file", command=donothing)
    menubar.add_cascade(label="view", menu=viewmenu)

   
    root.config(menu=menubar)




  






###############################################################################################################

#menu  window
def secwin(root):
    
    
    root.destroy()
    root1=Tk()
    photo=PhotoImage(file=bkimg2)
    Label(root1,image=photo).place(x=0,y=0)

    ajax(root1,'These are the main options to choose from')
  
    set_win(root1)
    menuinator(root1)
    
    
  
    def midspace():
        Label(f2,text= "              ",bg="blue").pack()

    f2=Frame(root1,bg=bkg,borderwidth=5,padx=20,pady=20,relief=SUNKEN)
    f2.place(x=450,y=30)
    
  


    B=Button(f2,text= "STOCK PRICES ",bg=brg,padx=30,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:stockwin(root1)).pack()
    midspace()
    B1=Button(f2,text="SELL  DETAILS ",bg=brg,padx=30,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:sellwin(root1)).pack()
    midspace()
    B2=Button(f2,text="ABOUT",padx='100',bg=brg,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN).pack()
    midspace()
    B3=Button(f2,text="HELP",padx='100',bg=brg,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN).pack()
    midspace()
    B5=Button(f2,text="QUIT",padx='100',bg=brg,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:home(root1)).pack()
    midspace()
    root1.mainloop()
    ##########################stay away , ded end##############################################






########################################################################################################################################\

# stock handling
def stockwin(root):
    root.destroy()

    root1=Tk()
    photo=PhotoImage(file=bkimg3)
    Label(root1,image=photo).place(x=0,y=0)
    
    set_win(root1)
  
    ajax(root1,'These are the main stock purchase handeling options ')
    menuinator(root1)
    def midspace():
         Label(f2,text= "              ",bg="blue").pack()
    f2=Frame(root1,bg=bkg,borderwidth=5,padx=20,pady=20,relief=SUNKEN)
    f2.place(x=450,y=30)
    B=Button(f2,text= "CREATE SHEET ",bg=brg,padx=30,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:data1.new_sheet()).pack()
    midspace()
    B1=Button(f2,text="EDIT SHEET ",bg=brg,padx=55,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:data1.edit_sheet()).pack()
    midspace()
    B4=Button(f2,text="QUIT",bg=brg,padx=120,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:home(root1)).pack()
    root1.mainloop()


def sellwin(root):
    root.destroy()
    root1=Tk()

    set_win(root1)
    photo=PhotoImage(file=bkimg4)
    Tphoto=Label(image=photo).place(x=100,y=0)
    
    menuinator(root1)





    def midspace():
         Label(f2,text= "              ",bg="blue").pack()
    f2=Frame(root1,bg=bkg,borderwidth=5,padx=20,pady=20,relief=SUNKEN)
    f2.place(x=450,y=30)
    B=Button(f2,text= "CREATE SHEET ",bg=brg,padx=30,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:data2.new_sheet()).pack()
    midspace()
    B1=Button(f2,text="EDIT SHEET ",bg=brg,padx=55,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:data2.edit_sheet()).pack()
    midspace()
    B4=Button(f2,text="QUIT",bg=brg,padx=120,fg=frg,font="elephant 19 bold",borderwidth=3,relief=SUNKEN,command=lambda:home(root1))
    B4.pack()
    root1.mainloop()


home()   
