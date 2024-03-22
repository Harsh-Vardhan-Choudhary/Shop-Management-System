import openpyxl as xl
from tkinter import*
from tkinter import messagebox
import os
import data2
from datetime import datetime
import sys



a=os.path.join(sys.path[0],'Records')
def edit_sheet(rootp='none'):  
        if rootp!='none':
            rootp.destroy()
      
        
        root1=Tk()
        root1.configure(background="white")
        root1.title("Shop Management")
        root1.geometry("{0}x{1}+0+0".format(root1.winfo_screenwidth(), root1.winfo_screenheight()))
        rootwidth=root1.winfo_screenwidth()
        rootheight=root1.winfo_screenheight()

    
        rnow=datetime.now()
        
        date=rnow.strftime("%d")
        month=rnow.strftime("%m")
        year=rnow.strftime("%Y")
        
    
        def feed_stock_data(sheet,wb,file_path):
            
            fx=Frame(root1,bg='grey',width=rootwidth,height=rootheight,border=8,relief=SUNKEN)
            fx.place(x=0,y=100)

            a=Label(fx,text='Feed DATA in Sheet',bg='red',fg='white',font="arial 11 bold")
            a.place(x=0,y=110)
            Label(fx,text='Name',fg='red',bg='grey',font="arial 19 bold",).place(x=0,y=130)
            Label(fx,text='Item',fg='red',bg='grey',font="arial 19 bold").place(x=0,y=170)
            Label(fx,text='Price/piece',fg='red',bg='grey',font="arial 19 bold").place(x=0,y=210)
            Label(fx,text='Quantity',fg='red',bg='grey',font="arial 19 bold").place(x=0,y=250)
            Label(fx,text='Contacts',fg='red',bg='grey',font="arial 19 bold").place(x=0,y=290)
            Label(fx,text=datetime.date(datetime.today()),fg='black',bg='grey',font="arial 11 bold",).place(x=0,y=30)
            Label(fx,text=datetime.time(datetime.now()),fg='black',bg='grey',font="arial 11 bold",).place(x=0,y=50)
        
    
            ITEM=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            ITEM.place(x=200,y=130)
            
            DET=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            DET.place(x=200,y=170)
            PRICE=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            PRICE.place(x=200,y=210)
            QUAN=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            QUAN.place(x=200,y=250)
            CONT=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            CONT.place(x=200,y=290)
         
            entries=[child for child in fx.winfo_children()
                     if isinstance(child,Entry)]

            def switch(Event,Lst,ths_index):
                next_index=(ths_index+1)%len(Lst)
                Lst[next_index].focus_set()
                
            for idx,entry in enumerate(entries):
                entry.bind('<Return>',lambda e,idx=idx:switch(e,entries,idx))
                
            
            
            def reciept(sheet):
                i=2   
                while sheet.cell(i,1).value:
                    i+=1   
                row=i-1
                name=sheet.cell(row,2).value
                
                
                
                while sheet.cell(row,2).value==name:
                     
                    row-=1
                x=messagebox.askquestion('Reciept',message=f"Do you want reciept of {name}")
                if x=='yes':
                    can(name,row,i)
            
            def can(name,row,i):
                c=Canvas(fx,bg='white',height=600,width=300)
                c.create_line(0,100,300,100,dash=(4,2))
                c.create_text(100,120,fill="darkblue",font="Times 15 italic bold",text=name)
                j=160
                row=row+1
                
                total=0
                while row<i:
                    data=sheet.cell(row,3).value+' : '+sheet.cell(row,4).value+' x '+sheet.cell(row,5).value
                    c.create_text(100,j,fill="darkblue",font="Times 15 italic bold",text=data)
                    total=int(sheet.cell(row,4).value)*int(sheet.cell(row,5).value)+total
                    row+=1
                    j+=50
                    ttl='total :-  Rs ',total
                c.create_line(0,j,300,j,dash=(4,2))
                c.create_text(100,j+30,fill="darkblue",font="Times 15 italic bold",text=ttl)
                c.create_rectangle(10,10,290,590)
                
                
                c.update()
                c.postscript(file='reciept.ps',colormode='color')
                c.place(x=900,y=10) 
            
            def yesnos(sheet):
                messagebox.askquestion('SAVING DATA',message='Are you sure,you want to save data')
                           
               
                i=2
                while sheet.cell(i,1).value:
                    i+=1   
                row=i   
                
                sheet.cell(row,1).value=row-1
                sheet.cell(row,2).value=ITEM.get()
                sheet.cell(row,3).value=DET.get()
                sheet.cell(row,4).value=PRICE.get() 
                sheet.cell(row,6).value=CONT.get() 
                sheet.cell(row,8).value=datetime.date(datetime.today()) 
                sheet.cell(row,5).value=QUAN.get()
                sheet.cell(row,7).value=datetime.time(datetime.now())
                wb.save(file_path)
                
            
                    
            Button(root1,text='SAVE',bg='green',fg='white',font="arial 19 bold",command=lambda:yesnos(sheet)).place(x=100,y=550)
            Button(root1,text='Receipt',bg='dark grey',fg='white',font="arial 19 bold",command=lambda:reciept(sheet)).place(x=250,y=550) 
            Button(root1,text='DISCARD',bg='red',fg='white',font="arial 19 bold",command=lambda:root1.destroy()).place(x=400,y=550)   
            Button(root1,text='BACK',bg='light blue',fg='white',font="arial 19 bold",command=lambda:root1.destroy()).place(x=600,y=550) 
            

            
             
        def  set_search_name():
           
            b=month+year
          
            try:
                file_path=a+'\\sell'+b+'.xlsx'
                wb=xl.load_workbook(file_path)
                sheet=wb['Sheet']
                feed_stock_data(sheet,wb,file_path)
                wb.save(file_path)
            except FileNotFoundError:
                           
                file_name=' '
                file_name='sell'+b+'.xlsx'
                wb=xl.Workbook()
                file_path=a+"\\"+file_name
                wb.save(file_path)
                sheet=wb['Sheet']
                data2.setsheet(sheet)
                feed_stock_data(sheet,wb,file_path)
                wb.save(file_path)
                
        
    
        set_search_name()

def show(rootp='none'):
    rnow=datetime.now()
    global a        
    date=rnow.strftime("%d")
    month=rnow.strftime("%m")
    year=rnow.strftime("%Y")
  
      
    b=month+year
          
    try:
                file_path=a+'\\sell'+b+'.xlsx'
                wb=xl.load_workbook(file_path)
                sheet=wb['Sheet']
                if rootp!='none':
                    rootp.destroy()
                
                root1=Toplevel()
                root1.configure(background="white")
                root1.geometry("{0}x{1}+0+0".format(root1.winfo_screenwidth(), root1.winfo_screenheight()))
                rootwidth=root1.winfo_screenwidth()
                rootheight=root1.winfo_screenheight()

                fx=Frame(root1,bg='grey',width=rootwidth,height=rootheight,border=8,relief=SUNKEN)
                fx.place(x=0,y=100)
                i=1
                while sheet.cell(i,1).value:
                    A=sheet.cell(i,1).value
                    b=sheet.cell(i,2).value
                    c=sheet.cell(i,3).value
                    d=sheet.cell(i,4).value
                    e=sheet.cell(i,5).value
                    f=sheet.cell(i,6).value
                    g=sheet.cell(i,7).value
                    h=sheet.cell(i,8).value
                    
                    Y=i*50
                    Label(fx,text=A,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=50,y=Y)
                    Label(fx,text=b,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=200,y=Y)
                    Label(fx,text=c,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=350,y=Y)
                    Label(fx,text=d,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=500,y=Y)
                    Label(fx,text=e,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=650,y=Y)
                    Label(fx,text=f,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=800,y=Y)
                    Label(fx,text=g,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=950,y=Y)
                    Label(fx,text=h,bg='grey',fg='white',font="arial 15 bold",relief=SUNKEN).place(x=1050,y=Y)
                    i+=1
                root1.mainloop()
            
            
            
            
             
    except FileNotFoundError:
                           
            messagebox.showerror(title='Sorry',message='File is empty')
              
              
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    