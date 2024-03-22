from tkinter import *
import trash1
import random
import openpyxl as xl
import data3
import os
from datetime import datetime
from tkinter import messagebox
import sys

OTP='none'

a=os.path.join(sys.path[0],'Records')
s=os.path.join(os.sys.path[0],'Records')




file_path=s+'\\settings.xlsx'
wb=xl.load_workbook(file_path)
sheet=wb['Sheet']
frg=sheet.cell(1,1).value
bkg=sheet.cell(2,1).value
brg=sheet.cell(3,1).value
bkimg=sheet.cell(4,1).value
bkimg2=sheet.cell(5,1).value
bkimg3=sheet.cell(6,1).value
bkimg4=sheet.cell(7,1).value
sound=sheet.cell(1,2).value

























def login_form(root='none'):
    if root!='none':
            root.destroy()
    
    root1=Toplevel()
    
    photo=PhotoImage(file=bkimg2)
    Tphoto=Label(root1,image=photo).pack()
    fz=Frame(root1,bg='blue',width=400,height=500,borderwidth=5,padx=20,pady=20,relief=SUNKEN)
    fz.place(relx='0.5',rely='0.5',anchor=CENTER)
    user=Entry(fz,fg='black',bg='white',font="arial 19 ")
    user.place(x=20,y=100)
    Label(fz,text='EMAIL ID',fg='white',bg='blue',font="arial 19 bold").place(x=20,y=50)
    passwrd=Entry(fz,fg='white',bg='white',font="arial 19 bold")
    passwrd.place(x=20,y=200)
    Label(fz,text='PASSWORD',fg='white',bg='blue',font="arial 19 bold").place(x=20,y=150)      
    switcher(fz)
    
         
    Button(fz,text="SUBMIT",fg="lime",font=("arial baltic",20),bg="blue",command=lambda:check()).place(x=60,y=400)
    def check():
        file_path=a+'\\credentials.xlsx'
        wb=xl.load_workbook(file_path)
        sheet=wb['Sheet']
        i=1
        found=False
        while sheet.cell(i,1).value:
            print(sheet.cell(i,2).value)
            if sheet.cell(i,2).value==user.get()and sheet.cell(i,3).value==passwrd.get():
                sheet.cell(i,7).value='logged in' 
                wb.save(file_path)           
                messagebox.showerror(title='logged in',message='Welcome')
                found=True
                root1.destroy()
            i+=1   
        if found==False:
            messagebox.showerror(title='Error',message=' User not found(either username or password is incorrect)')
    root1.mainloop()                   
    
    
    
def switcher(fz):    
    entries=[child for child in fz.winfo_children()
            if isinstance(child,Entry)]

    def switch(Event,Lst,ths_index):
        next_index=(ths_index+1)%len(Lst)
        Lst[next_index].focus_set()
                
    for idx,entry in enumerate(entries):
        entry.bind('<Return>',lambda e,idx=idx:switch(e,entries,idx))
        
        

def signin_form():
   
    root1=Toplevel()
    
    photo=PhotoImage(file=bkimg3)
    Tphoto=Label(root1,image=photo).pack()
    fz=Frame(root1,bg='blue',width=400,height=500,borderwidth=5,padx=20,pady=20,relief=SUNKEN)
    fz.place(relx='0.5',rely='0.5',anchor=CENTER)
    usrname=StringVar()
    user=Entry(fz,textvariable=usrname,fg='black',bg='white',font="arial 19 ")
    user.focus_set()
    user.place(x=20,y=100)
    Label(fz,text='User Name',fg='white',bg='blue',font="arial 19 bold").place(x=20,y=50)
    passwrd=Entry(fz,fg='grey',bg='white',font="arial 19 bold")
    passwrd.place(x=20,y=200)
    Label(fz,text='PASSWORD',fg='white',bg='blue',font="arial 19 bold").place(x=20,y=150) 
    
    passwrd=Entry(fz,fg='white',bg='grey',font="arial 19 bold")
    passwrd.place(x=20,y=300)
    Label(fz,text='CONFIRM PASSWORD',fg='white',bg='blue',font="arial 19 bold").place(x=20,y=250)     
    switcher(fz)
    Button(fz,text="NEXT",fg="lime",font=("arial baltic",20),bg="blue",command=lambda:toyo()).place(x=260,y=400)    
    def toyo():
        fz=Frame(root1,bg='blue',width=400,height=500,borderwidth=5,padx=20,pady=20,relief=SUNKEN)
    
        fz.place(relx='0.5',rely='0.5',anchor=CENTER)
        
        user2=Entry(fz,fg='black',bg='white',font="arial 19 ")
        user2.place(x=20,y=100)
        Label(fz,text='EMAIL ID',fg='white',bg='blue',font="arial 19 bold").place(x=20,y=50)
       
        Button(fz,text="SEND OTP",fg="lime",font=("arial baltic",15),bg="blue",command=lambda:send_otp()).place(x=200,y=150)   
        Checkbutton(fz,text='I am not a robot',fg='white',bg='blue',font="arial 10 bold").place(x=20,y=300)
       
        switcher(fz)
        def send_otp():
           
            a=str(random.randint(99,999))
            b=str(random.randint(99,999))
            OTP=a+b
            users=str(user2.get())
            
            print(users)
            try:
                trash1.mailboy('romanogers99@gmail',users,OTP)
            except SyntaxError:
                messagebox.showerror(title='warning',message='wrong id')                
            verify(OTP)
            
        
                   
               
            
            
            
        def verify(OTP):
           
            fz=Frame(root1,bg='red',width=400,height=500,borderwidth=5,padx=20,pady=20,relief=SUNKEN)
            fz.place(relx='0.5',rely='0.5',anchor=CENTER)   
    
            otp=Entry(fz,fg='Black',bg='white',font="arial 19 bold")
            otp.focus_set()
            otp.place(x=20,y=200)
            def ver(OTP):
                if otp.get()==OTP:
                    
                    Button(fz,text="SUBMIT",fg="lime",font=("arial baltic",20),bg="blue",command=lambda:feed_cred()).place(x=200,y=350) 
                    return 1
                else :
                    pass          
            Button(fz,text="Verify",fg="lime",font=("arial baltic",20),bg="blue",command=lambda:ver(OTP)).place(x=260,y=250)
            
            
            
           
                   
            
         
        
        def setsheet(sheet):
                sheet.cell(1,1).value='S.no'
                sheet.cell(1,2).value='USERName'
                sheet.cell(1,3).value='PASSWORD'
                sheet.cell(1,4).value='EMAIL ID'
                sheet.cell(1,5).value='date'
                sheet.cell(1,6).value='verified'
        def feed_cred():
                 
            try:
                file_path=a+'\\credentials.xlsx'
                wb=xl.load_workbook(file_path)
                sheet=wb['Sheet']
                #feed_cred(sheet,wb,file_path)
                setsheet(sheet)
                wb.save(file_path)
                 
                i=2
                while sheet.cell(i,1).value:
                    i+=1   
                row=i   
                
                sheet.cell(row,1).value=row-1
                sheet.cell(row,2).value=user.get()
                sheet.cell(row,3).value=passwrd.get()
                sheet.cell(row,4).value=user2.get()
                sheet.cell(row,6).value='true'
                sheet.cell(row,5).value=datetime.date(datetime.now()) 
                
                wb.save(file_path)
                messagebox.showerror(title='Welecome',message='Congratulations!!! You are registered')
                root1.destroy()
                
                
               
            except FileNotFoundError:
                           
                file_name=' '
                file_name='credentials.xlsx'
                wb=xl.Workbook()
                file_path=a+"\\"+file_name
                wb.save(file_path)
                sheet=wb['Sheet']
                setsheet(sheet)
                feed_cred()
                wb.save(file_path)       
    root1.mainloop()

def logout():
    try:
                file_path=a+'\\credentials.xlsx'
                wb=xl.load_workbook(file_path)
                sheet=wb['Sheet']
                wb.save(file_path)
                 
                i=1
                while sheet.cell(i,1).value:
                    if sheet.cell(i,7).value=='logged in':
                        sheet.cell(i,7).value=''
                        wb.save(file_path)
                        messagebox.showerror(title='Welecome',message='Logged out')
                        break
                    else:
                        i+=1   
                        
                
                
                
               
    except FileNotFoundError:
                messagebox.showerror(title='Welecome',message='Sign in Please')
 
 

