import openpyxl as xl
from tkinter import*
from tkinter import messagebox
import os
import settings
import sys

a=os.path.join(sys.path[0],'Records')

s=os.path.join(sys.path[0],'Records')
settings.deftheme()
file_path=s+'\\settings.xlsx'
wb=xl.load_workbook(file_path)
sheet=wb['Sheet']
frg=sheet.cell(1,1).value
bkg=sheet.cell(2,1).value
brg=sheet.cell(3,1).value
bkimg=sheet.cell(4,1).value
bkimg2=sheet.cell(5,1).value
bkimg3=sheet.cell(6,1).value
bkimg4=sheet.cell(7,1).value





def new_sheet():   
        
        root1=Tk()
        root1.configure(background=bkg)
        root1.title("Shopiy")
        root1.geometry("300x200")
        root1.maxsize(500,200)
        root1.minsize(500,200)
       
        
            
        Label(root1,text='CREATE NEW SHEET',bg=bkg,padx=50,fg=frg,font="arial 11 bold").grid(row=1,column=1)
        Label(root1,text='MONTH ',bg=bkg,padx=10,fg="black",font="arial 11 bold").grid(row=3,column=1)
        Label(root1,text='YEAR ',bg=bkg,padx=10,fg="black",font="arial 11 bold").grid(row=5,column=1)
        
        #feed month and year
        MONTHENTRY=Entry(root1,bg="white",fg=frg,font="arial 11 bold",borderwidth=1,relief=SUNKEN)
        MONTHENTRY.grid(row=3,column=2)

        YEARENTRY=Entry(root1,bg="white",fg=frg,font="arial 11 bold",borderwidth=1,relief=SUNKEN)
        YEARENTRY.grid(row=5,column=2)
        
        
        
        
        Label(root1).grid(row=6,column=1)
        Button(root1,text='CREATE',bg='green',fg='white',command=lambda:new_sheet2()) .grid(row=7,column=2)
        Button(root1,text='DISCARD',bg='red',fg='white',command=lambda:pass2(root1)).grid(row=7,column=1)
        
        def pass2(root):
            root.destroy()

        def new_sheet2():
            Label(root1,bg='white').grid(row=8,column=1)
            Label(root1,bg='white').grid(row=9,column=1)
            
            month_name=MONTHENTRY.get()
            month_name=month_name[:3].upper()
            years_name=YEARENTRY.get()
            years_name=years_name[:4]
            
            file_name=' '
            file_name='stock'+month_name+years_name+'.xlsx'
            wb=xl.Workbook()
            file_path=a+"\\"+file_name
            wb.save(file_path)
            sheet=wb['Sheet']
            setsheet(sheet)
            wb.save(file_path)
            root1.destroy()
            #Label(root1,text='created ',bg="red",padx=50,fg="black",font="arial 11 bold").grid(row=10,column=1,columnspan=3)




def edit_sheet(rootp='none'):
        if rootp!='none':
            rootp.destroy()
        root1=Toplevel()
        photo=PhotoImage(file=bkimg2)
        Label(root1,image=photo).place(x=0,y=0)

        root1.title("Shopify")
        root1.geometry("{0}x{1}+0+0".format(root1.winfo_screenwidth(), root1.winfo_screenheight()))
        rootwidth=root1.winfo_screenwidth()
        rootheight=root1.winfo_screenheight()


        search=Entry(root1,bg=bkg,fg=frg,font="arial 19 bold",width=rootwidth,borderwidth=6,relief=GROOVE)
        search.place(x=100,y=10)
    
     
        
        def  set_search_name():
            b0=search.get()
            b1=b0[:3].upper()
            b2=b0[3:7]
            b=b1+b2

            try:
                file_path=a+'\\stock'+b+'.xlsx'
                wb=xl.load_workbook(file_path)
                sheet=wb['Sheet']
                feed_stock_data(sheet,wb,file_path)
                wb.save(file_path)
            except FileNotFoundError:
                messagebox.showerror(title='ERROR 404',message='FILE NOT FOUND')
               
        
        B=Button(root1,text='search',bg=brg,font="arial 11 bold",relief=RAISED,command=lambda:set_search_name())
        B.place(x=20,y=10)

                

        def feed_stock_data(sheet,wb,file_path):
            
            fx=Frame(root1,bg=brg,width=rootwidth,height=rootheight,border=8,relief=SUNKEN)
            fx.place(x=0,y=100)
            photo=PhotoImage(file=bkimg2)
            Label(fx,image=photo).place(x=0,y=0)

            a=Label(fx,text='Feed DATA in Sheet',bg=bkg,fg=frg,font="arial 11 bold")
            a.place(x=0,y=110)
            Label(fx,text='Item Name',fg=frg,bg=bkg,font="arial 19 bold",).place(x=0,y=130)
            Label(fx,text='Detail',fg=frg,bg=bkg,font="arial 19 bold").place(x=0,y=170)
            Label(fx,text='Price/piece',fg=frg,bg=bkg,font="arial 19 bold").place(x=0,y=210)
            Label(fx,text='Quantity',fg=frg,bg=bkg,font="arial 19 bold").place(x=0,y=250)

           
            ITEM=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            ITEM.place(x=200,y=130)
            
            DET=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            DET.place(x=200,y=170)
            PRICE=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            PRICE.place(x=200,y=210)
            QUAN=Entry(fx,fg='red',bg='white',font="arial 19 bold")
            QUAN.place(x=200,y=250)
         
            
            entries=[child for child in fx.winfo_children()
                    if isinstance(child,Entry)
                    ]         
            def go_to_next_entry(Event,entry_list,this_index):
                next_index=(this_index+1)%len(entry_list)
                entry_list[next_index].focus_set()
                
            for idx, entry in enumerate(entries):
                entry.bind('<Return>', lambda e, idx=idx: go_to_next_entry(e, entries, idx))
            
           
          
            def yesnos(sheet):
                messagebox.askquestion('SAVING DATA',message='Are you sure,you want to save data')
                           
               
                i=2
               
                while sheet.cell(i,1).value:
                    i+=1   
                row=i   
                
                sheet.cell(row,1).value=row-1
                sheet.cell(row,2).value=ITEM.get()
                sheet.cell(row,3).value=DET.get()
                sheet.cell(row,4).value=PRICE.get()
                sheet.cell(row,5).value=QUAN.get() 
                wb.save(file_path)
          
                
            Button(root1,text='SAVE',bg='green',fg='white',font="arial 19 bold",command=lambda:yesnos(sheet)).place(x=300,y=550)
            Button(root1,text='DISCARD',bg='red',fg='white',font="arial 19 bold",command=lambda :root1.destroy()).place(x=600,y=550)   
            
            

           
              
              
        






def setsheet(sheet):
    sheet.cell(1,1).value='S.no'
    sheet.cell(1,2).value='Item name'
    sheet.cell(1,3).value='details'
    sheet.cell(1,4).value='price/piece'
    sheet.cell(1,5).value='quantity'


